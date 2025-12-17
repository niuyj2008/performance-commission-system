#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
import sys

def analyze_excel(file_path):
    print(f"\n分析文件: {file_path}")
    print("=" * 80)
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        print(f"\n工作表列表: {wb.sheetnames}")
        
        # 分析每个工作表
        for sheet_name in wb.sheetnames[:3]:  # 只看前3个工作表
            print(f"\n--- 工作表: {sheet_name} ---")
            ws = wb[sheet_name]
            
            # 显示前10行数据
            print(f"维度: {ws.max_row} 行 x {ws.max_column} 列")
            print("\n前10行数据:")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                print(f"第{row_idx}行: {row}")
            
            # 查找包含公式的单元格
            print("\n包含公式的单元格示例（前10个）:")
            formula_count = 0
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        print(f"  {cell.coordinate}: {cell.value}")
                        formula_count += 1
                        if formula_count >= 10:
                            break
                if formula_count >= 10:
                    break
            
            if formula_count == 0:
                print("  未找到公式")
                
    except Exception as e:
        print(f"错误: {e}")

if __name__ == "__main__":
    # 分析主要文件
    files = [
        "/Users/tima/Documents/ZhangHR/2025上半年/2025年2月至7月设计奖.xlsx",
        "/Users/tima/Documents/ZhangHR/2025上半年/空调面积表/建业大厦项目空调面积表+.xlsm"
    ]
    
    for file in files:
        analyze_excel(file)
