#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

file_path = "/Users/tima/Documents/ZhangHR/2025上半年/ALL奖金发放表2025年7月8日含预支（20250716).xlsx"
print(f"\n分析文件: {file_path}")
print("=" * 80)

wb = openpyxl.load_workbook(file_path, data_only=False)
print(f"\n工作表列表: {wb.sheetnames}\n")

# 查看每个工作表
for sheet_name in wb.sheetnames[:5]:  # 前5个工作表
    print(f"\n{'='*60}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*60}")
    ws = wb[sheet_name]
    print(f"维度: {ws.max_row} 行 x {ws.max_column} 列\n")
    
    # 显示前20行
    print("前20行数据:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        # 过滤掉全是None的行
        if any(cell is not None for cell in row):
            print(f"第{row_idx}行: {row}")
    
    # 查找公式
    print("\n包含公式的单元格（前15个）:")
    formula_count = 0
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"  {cell.coordinate}: {cell.value}")
                formula_count += 1
                if formula_count >= 15:
                    break
        if formula_count >= 15:
            break
    
    if formula_count == 0:
        print("  未找到公式")
