#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl

def analyze_project_file(file_path):
    print(f"\n分析文件: {file_path}")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print(f"\n工作表列表: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames[:2]:
            print(f"\n--- 工作表: {sheet_name} ---")
            ws = wb[sheet_name]
            print(f"维度: {ws.max_row} 行 x {ws.max_column} 列")
            
            # 显示前15行
            print("\n前15行数据:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
                print(f"第{row_idx}行: {row}")
            
            # 查找公式
            print("\n包含公式的单元格示例（前20个）:")
            formula_count = 0
            for row in ws.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        print(f"  {cell.coordinate}: {cell.value}")
                        formula_count += 1
                        if formula_count >= 20:
                            break
                if formula_count >= 20:
                    break
                    
    except Exception as e:
        print(f"错误: {e}")

# 分析项目清单
analyze_project_file("/Users/tima/Documents/ZhangHR/2025上半年/2025年项目上半年清单.xlsx")
